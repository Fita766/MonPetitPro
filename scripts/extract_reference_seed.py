"""Generate the MonPetitPro reference seed from official and client sources.

The generated SQL contains only reference values and communes. It never
imports operation rows from the workbook.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


DEPARTMENTS = ("02", "59", "60", "62", "77", "80", "93", "94", "95")
DEPARTMENT_NAMES = {
    "02": "Aisne",
    "59": "Nord",
    "60": "Oise",
    "62": "Pas-de-Calais",
    "77": "Seine-et-Marne",
    "80": "Somme",
    "93": "Seine-Saint-Denis",
    "94": "Val-de-Marne",
    "95": "Val-d'Oise",
}
REGION_NAMES = {
    "02": "Hauts-de-France",
    "59": "Hauts-de-France",
    "60": "Hauts-de-France",
    "62": "Hauts-de-France",
    "80": "Hauts-de-France",
    "77": "Île-de-France",
    "93": "Île-de-France",
    "94": "Île-de-France",
    "95": "Île-de-France",
}
ZONE_URL = (
    "https://static.data.gouv.fr/resources/liste-des-communes-selon-le-zonage-abc/"
    "20260703-091314/"
    "liste-ensemble-des-communes-zonage-abc-en-vigueur-26-juin-2026.csv"
)
GEO_URL = (
    "https://geo.api.gouv.fr/departements/{department}/communes"
    "?fields=nom,code,codesPostaux,codeDepartement,codeRegion&format=json"
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def sql_text(value: str | None) -> str:
    if value is None or value == "":
        return "null"
    return "'" + value.replace("'", "''") + "'"


def fetch_text(url: str) -> str:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "MonPetitPro reference seed/1.0"},
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read().decode("utf-8-sig")


def fetch_json(url: str) -> Any:
    return json.loads(fetch_text(url))


def read_zones(zone_file: Path | None) -> dict[str, str]:
    content = (
        zone_file.read_text(encoding="utf-8-sig")
        if zone_file
        else fetch_text(ZONE_URL)
    )
    rows = csv.DictReader(io.StringIO(content), delimiter=";")
    zone_column = "Zonage ABC en vigueur depuis le 26 juin 2026"
    return {
        clean(row.get("CODGEO")): clean(row.get(zone_column))
        for row in rows
        if clean(row.get("DEP")) in DEPARTMENTS
    }


def read_communes(
    zones: dict[str, str],
    geo_file: Path | None,
) -> list[dict[str, str | None]]:
    communes: list[dict[str, str | None]] = []
    offline_rows: dict[str, list[dict[str, Any]]] = {}
    if geo_file:
        loaded = json.loads(geo_file.read_text(encoding="utf-8-sig"))
        for department, rows in loaded.items():
            if department not in DEPARTMENTS:
                continue
            if (
                isinstance(rows, list)
                and len(rows) == 1
                and isinstance(rows[0], dict)
                and isinstance(rows[0].get("value"), list)
            ):
                rows = rows[0]["value"]
            elif isinstance(rows, dict) and isinstance(rows.get("value"), list):
                rows = rows["value"]
            if not isinstance(rows, list):
                raise ValueError(
                    f"Format de communes invalide pour le département {department}"
                )
            offline_rows[department] = rows
    for department in DEPARTMENTS:
        rows = (
            offline_rows.get(department, [])
            if geo_file
            else fetch_json(GEO_URL.format(department=department))
        )
        for row in rows:
            insee_code = clean(row.get("code"))
            postal_codes = row.get("codesPostaux") or []
            communes.append(
                {
                    "name": clean(row.get("nom")).upper(),
                    "insee_code": insee_code,
                    "postal_code": clean(postal_codes[0]) if postal_codes else None,
                    "department_code": department,
                    "department_name": DEPARTMENT_NAMES[department],
                    "region_name": REGION_NAMES[department],
                    "housing_zone": zones.get(insee_code) or None,
                }
            )
    return sorted(communes, key=lambda row: (str(row["department_code"]), str(row["name"])))


def add(values: dict[str, set[str]], kind: str, value: Any) -> None:
    label = clean(value)
    if label and not label.startswith("#"):
        values[kind].add(label)


def read_workbook_references(workbook_path: Path) -> dict[str, set[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    values: dict[str, set[str]] = defaultdict(set)

    bd = workbook["BD"]
    bd_columns = {
        "C": "certification",
        "E": "thermal_regulation",
        "L": "cop",
        "N": "ctx",
        "P": "promoter",
    }
    for column, kind in bd_columns.items():
        column_index = column_index_from_string(column)
        for row in bd.iter_rows(
            min_row=4,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            add(values, kind, row[0])

    board = workbook["TBL BORD"]
    board_columns = {
        21: "promoter",
        22: "cop",
        23: "ctx",
        24: "assistant",
        25: "gpa_assistant",
        34: "manager",
        35: "animation_provider",
    }
    for row in board.iter_rows(min_row=4, max_col=35, values_only=True):
        for one_based_column, kind in board_columns.items():
            add(values, kind, row[one_based_column - 1])

    values["program_nature"].update(
        {
            "Neuf",
            "Réhabilitation",
            "Démolition",
            "Étudiant",
            "Béguinage",
            "Commerce",
            "Mixte",
            "Autre",
        }
    )
    return values


def value_batches(rows: list[str], size: int = 400) -> Iterable[list[str]]:
    for index in range(0, len(rows), size):
        yield rows[index : index + size]


def render_seed(
    references: dict[str, set[str]],
    communes: list[dict[str, str | None]],
) -> str:
    output = [
        "-- Généré par scripts/extract_reference_seed.py.",
        "-- Sources : classeur DMO, API Découpage administratif et zonage ABC 2026.",
        "begin;",
        "",
    ]

    reference_rows: list[str] = []
    for kind in sorted(references):
        for order, label in enumerate(sorted(references[kind], key=str.casefold), start=1):
            reference_rows.append(
                f"({sql_text(kind)}, {sql_text(label)}, {order})"
            )
    for batch in value_batches(reference_rows):
        output.extend(
            [
                "insert into public.reference_values (kind, label, sort_order)",
                "values",
                ",\n".join(batch),
                "on conflict (kind, normalized_label) do update",
                "set label = excluded.label,",
                "    is_active = true,",
                "    sort_order = excluded.sort_order;",
                "",
            ]
        )

    commune_rows = [
        "("
        + ", ".join(
            [
                sql_text(str(row["name"])),
                sql_text(str(row["insee_code"])),
                sql_text(row["postal_code"]),
                sql_text(str(row["department_code"])),
                sql_text(str(row["department_name"])),
                sql_text(str(row["region_name"])),
                sql_text(row["housing_zone"]),
            ]
        )
        + ")"
        for row in communes
    ]
    for batch in value_batches(commune_rows):
        output.extend(
            [
                "insert into public.communes (",
                "  name, insee_code, postal_code, department_code,",
                "  department_name, region_name, housing_zone",
                ") values",
                ",\n".join(batch),
                "on conflict (insee_code) do update",
                "set name = excluded.name,",
                "    postal_code = excluded.postal_code,",
                "    department_code = excluded.department_code,",
                "    department_name = excluded.department_name,",
                "    region_name = excluded.region_name,",
                "    housing_zone = excluded.housing_zone,",
                "    is_active = true;",
                "",
            ]
        )

    output.extend(
        [
            "do $$",
            "declare missing_zones integer;",
            "begin",
            "  select count(*) into missing_zones",
            "  from public.communes",
            "  where department_code = any(array['02','59','60','62','77','80','93','94','95'])",
            "    and housing_zone is null;",
            "  if missing_zones > 0 then",
            "    raise exception 'référentiel communes incomplet : % zonage(s) manquant(s)', missing_zones;",
            "  end if;",
            "end $$;",
            "",
            "commit;",
            "",
        ]
    )
    return "\n".join(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--zone-file", type=Path)
    parser.add_argument("--geo-file", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.workbook.exists():
        print(f"Classeur introuvable : {args.workbook}", file=sys.stderr)
        return 2
    zones = read_zones(args.zone_file)
    communes = read_communes(zones, args.geo_file)
    references = read_workbook_references(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(render_seed(references, communes), encoding="utf-8")
    print(
        f"{len(communes)} communes et "
        f"{sum(len(items) for items in references.values())} références générées"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
